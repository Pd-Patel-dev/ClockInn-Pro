"""
Cash Drawer Export Service

Generates PDF and Excel exports for cash drawer sessions.
"""
from typing import List
from uuid import UUID
from datetime import date, datetime
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
import re

from app.models.cash_drawer import CashDrawerSession
from app.models.user import User
from app.models.company import Company
from app.services.company_service import get_company_settings


def sanitize_html(text: str) -> str:
    """Sanitize text for ReportLab Paragraph."""
    if not text:
        return ""
    text = str(text)
    # Remove script tags
    text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.IGNORECASE | re.DOTALL)
    # Escape & if not already an entity
    text = re.sub(r'&(?!\w+;)', '&amp;', text)
    # Remove path-like characters that might confuse reportlab (but keep / for dates)
    text = re.sub(r'[<>:"|?*\\]', '', text)
    # Ensure no leading/trailing spaces that could cause issues
    text = text.strip()
    return text


async def generate_cash_drawer_pdf(
    db: AsyncSession,
    company_id: UUID,
    sessions: List[CashDrawerSession],
    from_date: date,
    to_date: date,
) -> BytesIO:
    """Generate PDF report for cash drawer sessions."""
    import logging
    logger = logging.getLogger(__name__)
    
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    # Get company info
    result = await db.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    company_name = str(company.name) if company and company.name else "Company"
    company_name = re.sub(r'[<>:"|?*\\]', '', company_name).strip() or "Company"
    
    # Create buffer and document
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
    )
    story = []
    styles = getSampleStyleSheet()
    
    # Simple colors
    header_bg = colors.HexColor('#374151')
    
    # Simple header
    title_style = ParagraphStyle('Title', fontSize=14, fontName='Helvetica-Bold')
    subtitle_style = ParagraphStyle('Subtitle', fontSize=9, textColor=colors.gray)
    
    period_str = f"{from_date.strftime('%m/%d/%Y')} - {to_date.strftime('%m/%d/%Y')}"
    
    story.append(Paragraph(f"{company_name} - Cash Drawer Report", title_style))
    story.append(Paragraph(f"Period: {period_str} | Generated: {datetime.utcnow().strftime('%m/%d/%Y %I:%M %p')}", subtitle_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Simple table with fewer columns
    if sessions:
        # Header row - simple text, no Paragraph objects
        table_data = [["Date", "Employee", "Start", "End", "+/-", "Status"]]
        
        total_delta = 0
        total_start = 0
        total_end = 0
        
        for session in sessions:
            # Get employee name
            emp_result = await db.execute(select(User).where(User.id == session.employee_id))
            employee = emp_result.scalar_one_or_none()
            emp_name = str(employee.name)[:15] if employee and employee.name else "Unknown"
            
            date_str = session.start_counted_at.strftime("%m/%d/%y")
            start_cash = f"${session.start_cash_cents / 100:.0f}"
            end_cash = f"${session.end_cash_cents / 100:.0f}" if session.end_cash_cents else "-"
            
            delta_val = session.delta_cents or 0
            if delta_val > 0:
                delta_str = f"+${delta_val / 100:.0f}"
            elif delta_val < 0:
                delta_str = f"-${abs(delta_val) / 100:.0f}"
            else:
                delta_str = "$0"
            
            status = "Open" if session.status.value == "OPEN" else "OK" if session.status.value == "CLOSED" else "Review"
            
            total_delta += delta_val
            total_start += session.start_cash_cents
            if session.end_cash_cents:
                total_end += session.end_cash_cents
            
            table_data.append([date_str, emp_name, start_cash, end_cash, delta_str, status])
        
        # Totals row
        total_delta_str = f"+${total_delta / 100:.0f}" if total_delta >= 0 else f"-${abs(total_delta) / 100:.0f}"
        table_data.append(["TOTAL", f"{len(sessions)} sessions", f"${total_start / 100:.0f}", f"${total_end / 100:.0f}", total_delta_str, ""])
        
        # Create compact table
        col_widths = [0.8*inch, 1.5*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.6*inch]
        table = Table(table_data, colWidths=col_widths)
        
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (2, 1), (4, -1), 'RIGHT'),
            ('ALIGN', (5, 1), (5, -1), 'CENTER'),
            
            # Totals row
            ('BACKGROUND', (0, -1), (-1, -1), header_bg),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            
            # Grid and padding
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        story.append(table)
    else:
        story.append(Paragraph("No cash drawer sessions found for this period", styles['Normal']))
    
    # Build PDF
    try:
        doc.build(story)
        buffer.seek(0)
        return buffer
    except Exception as e:
        logger.error(f"Error building PDF: {str(e)}", exc_info=True)
        raise ValueError(f"Failed to generate PDF: {str(e)}")


async def generate_cash_drawer_excel(
    db: AsyncSession,
    company_id: UUID,
    sessions: List[CashDrawerSession],
    from_date: date,
    to_date: date,
) -> BytesIO:
    """Generate Excel report for cash drawer sessions."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Drawer"
    
    # Simple header
    headers = ["Date", "Employee", "Start", "End", "+/-", "Status"]
    ws.append(headers)
    
    # Style header
    header_fill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=9)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for session in sessions:
        emp_result = await db.execute(select(User).where(User.id == session.employee_id))
        employee = emp_result.scalar_one_or_none()
        emp_name = employee.name if employee else "Unknown"
        
        status = "Open" if session.status.value == "OPEN" else "OK" if session.status.value == "CLOSED" else "Review"
        
        ws.append([
            session.start_counted_at.strftime("%m/%d/%y"),
            emp_name,
            session.start_cash_cents / 100,
            session.end_cash_cents / 100 if session.end_cash_cents else None,
            session.delta_cents / 100 if session.delta_cents else 0,
            status,
        ])
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
