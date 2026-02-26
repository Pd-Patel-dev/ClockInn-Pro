from typing import List
from uuid import UUID
from datetime import datetime, date, timedelta
from io import BytesIO
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from app.models.time_entry import TimeEntry, TimeEntryStatus
from app.models.user import User


class NumberedCanvas:
    """Custom canvas for page numbers and headers/footers."""
    def __init__(self, canvas, doc):
        self.canvas = canvas
        self.doc = doc
        
    def draw_page_number(self, page_num):
        """Draw page number at bottom center."""
        self.canvas.saveState()
        self.canvas.setFont("Helvetica", 9)
        self.canvas.setFillColor(colors.HexColor('#666666'))
        page_text = f"Page {page_num}"
        text_width = self.canvas.stringWidth(page_text, "Helvetica", 9)
        page_width = self.doc.pagesize[0]
        self.canvas.drawString((page_width - text_width) / 2, 0.5 * inch, page_text)
        self.canvas.restoreState()


async def generate_pdf_report(
    db: AsyncSession,
    company_id: UUID,
    employee_ids: List[UUID],
    start_date: date,
    end_date: date,
    generated_by: str = "System",
) -> BytesIO:
    """Generate professional PDF report for time entries - one page per employee."""
    from app.pdf_templates.time_attendance_report import generate_time_attendance_report_pdf
    from app.models.company import Company
    from app.models.time_entry import TimeEntry, TimeEntryStatus
    from app.services.rounding_service import (
        compute_minutes_with_rounding_and_breaks,
        get_company_rounding_policy,
    )
    from app.services.company_service import get_company_settings
    
    # Get company information
    result = await db.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    company_name = company.name if company else "Company"
    
    # Get company settings for rounding and timezone
    if company:
        company_settings = get_company_settings(company)
        rounding_policy = company_settings["rounding_policy"]
        breaks_paid = company_settings["breaks_paid"]
        timezone_str = company_settings.get("timezone", "America/Chicago")
    else:
        rounding_policy = await get_company_rounding_policy(db, company_id)
        breaks_paid = False
        timezone_str = "America/Chicago"
    
    # Import timezone conversion functions
    from app.services.timezone_service import convert_to_company_timezone, get_utc_range_for_company_date_range

    # UTC bounds for the date range in company timezone (so export includes correct days)
    start_utc, end_utc = get_utc_range_for_company_date_range(timezone_str, start_date, end_date)

    # Get employees
    result = await db.execute(
        select(User).where(
            and_(
                User.id.in_(employee_ids),
                User.company_id == company_id,
            )
        ).order_by(User.name)
    )
    employees = result.scalars().all()
    
    # Prepare employee data for template
    employees_data = []
    
    for employee in employees:
        # Get time entries (filter by UTC range for company date range)
        result = await db.execute(
            select(TimeEntry).where(
                and_(
                    TimeEntry.employee_id == employee.id,
                    TimeEntry.company_id == company_id,
                    TimeEntry.clock_in_at >= start_utc,
                    TimeEntry.clock_in_at <= end_utc,
                )
            ).order_by(TimeEntry.clock_in_at)
        )
        entries = result.scalars().all()
        
        # Calculate totals
        total_hours = 0.0
        total_break_minutes = 0
        entry_count = 0
        entry_list = []
        
        for entry in entries:
            entry_count += 1
            # Convert UTC times to company timezone
            clock_in_local = convert_to_company_timezone(entry.clock_in_at, timezone_str)
            clock_out_local = convert_to_company_timezone(entry.clock_out_at, timezone_str) if entry.clock_out_at else None
            
            date_str = clock_in_local.strftime("%a, %m/%d/%Y")
            clock_in = clock_in_local.strftime("%I:%M %p")
            clock_out = clock_out_local.strftime("%I:%M %p") if clock_out_local else "Open"
            
            if entry.clock_out_at:
                rounded_minutes = compute_minutes_with_rounding_and_breaks(
                    entry.clock_in_at,
                    entry.clock_out_at,
                    entry.break_minutes,
                    rounding_policy,
                    breaks_paid,
                )
                hours = rounded_minutes / 60.0
                total_hours += hours
            else:
                hours = 0.0
            
            total_break_minutes += entry.break_minutes
            
            entry_list.append({
                'date': date_str,
                'clock_in': clock_in,
                'clock_out': clock_out,
                'hours': hours,
                'break_minutes': entry.break_minutes,
                'status': entry.status.value.title(),
            })
        
        # Calculate average hours per day
        avg_hours_per_day = (total_hours / entry_count) if entry_count > 0 else 0.0
        
        employees_data.append({
            'employee_name': employee.name,
            'job_role': employee.job_role or '',
            'total_entries': entry_count,
            'total_hours': total_hours,
            'total_break_minutes': total_break_minutes,
            'avg_hours_per_day': avg_hours_per_day,
            'entries': entry_list,
        })
    
    # Generate PDF using new template
    pdf_bytes = generate_time_attendance_report_pdf(
        company_name=company_name,
        period_start=start_date,
        period_end=end_date,
        generated_at=datetime.now(),
        generated_by=generated_by,
        employees_data=employees_data,
    )
    
    # Convert bytes to BytesIO
    buffer = BytesIO(pdf_bytes)
    buffer.seek(0)
    return buffer


async def generate_excel_report(
    db: AsyncSession,
    company_id: UUID,
    employee_ids: List[UUID],
    start_date: date,
    end_date: date,
) -> BytesIO:
    """Generate Excel report for time entries."""
    from app.models.company import Company
    from app.services.company_service import get_company_settings
    from app.services.timezone_service import convert_to_company_timezone, get_utc_range_for_company_date_range

    # Get company timezone
    result = await db.execute(select(Company).where(Company.id == company_id))
    company = result.scalar_one_or_none()
    if company:
        company_settings = get_company_settings(company)
        timezone_str = company_settings.get("timezone", "America/Chicago")
    else:
        timezone_str = "America/Chicago"

    # UTC bounds for the date range in company timezone
    start_utc, end_utc = get_utc_range_for_company_date_range(timezone_str, start_date, end_date)
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_headers = ['Employee', 'Total Hours', 'Regular Hours', 'Overtime Hours', 'Total Break Minutes']
    summary_ws.append(summary_headers)
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Detailed sheet
    detail_ws = wb.create_sheet("Detailed")
    detail_headers = ['Employee', 'Date', 'Clock In', 'Clock Out', 'Hours', 'Break (min)', 'Status']
    detail_ws.append(detail_headers)
    
    for cell in detail_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # UTC bounds for the date range in company timezone
    start_utc, end_utc = get_utc_range_for_company_date_range(timezone_str, start_date, end_date)

    # Get employees
    result = await db.execute(
        select(User).where(
            and_(
                User.id.in_(employee_ids),
                User.company_id == company_id,
            )
        )
    )
    employees = result.scalars().all()
    
    for employee in employees:
        # Get time entries (filter by UTC range for company date range)
        result = await db.execute(
            select(TimeEntry).where(
                and_(
                    TimeEntry.employee_id == employee.id,
                    TimeEntry.company_id == company_id,
                    TimeEntry.clock_in_at >= start_utc,
                    TimeEntry.clock_in_at <= end_utc,
                )
            ).order_by(TimeEntry.clock_in_at)
        )
        entries = result.scalars().all()
        
        if not entries:
            continue
        
        # Calculate totals
        total_hours = 0
        total_break_minutes = 0
        
        for entry in entries:
            if entry.clock_out_at:
                # Use rounding service for consistent calculation
                from app.services.rounding_service import (
                    compute_minutes_with_rounding_and_breaks,
                    get_company_rounding_policy,
                )
                from app.services.company_service import get_company_settings
                from app.models.company import Company
                
                # Get company settings
                result = await db.execute(
                    select(Company).where(Company.id == company_id)
                )
                company = result.scalar_one_or_none()
                if company:
                    company_settings = get_company_settings(company)
                    rounding_policy = company_settings["rounding_policy"]
                    breaks_paid = company_settings["breaks_paid"]
                else:
                    rounding_policy = await get_company_rounding_policy(db, company_id)
                    breaks_paid = False
                
                rounded_minutes = compute_minutes_with_rounding_and_breaks(
                    entry.clock_in_at,
                    entry.clock_out_at,
                    entry.break_minutes,
                    rounding_policy,
                    breaks_paid,
                )
                hours = rounded_minutes / 60.0
                total_hours += hours
            total_break_minutes += entry.break_minutes
            
            # Convert UTC times to company timezone for display
            clock_in_local = convert_to_company_timezone(entry.clock_in_at, timezone_str)
            clock_out_local = convert_to_company_timezone(entry.clock_out_at, timezone_str) if entry.clock_out_at else None
            
            # Add to detailed sheet (use rounded hours)
            detail_ws.append([
                employee.name,
                clock_in_local.strftime("%a, %Y-%m-%d"),
                clock_in_local.strftime("%H:%M"),
                clock_out_local.strftime("%H:%M") if clock_out_local else "Open",
                f"{hours:.2f}" if entry.clock_out_at else "0.00",
                entry.break_minutes,
                entry.status.value,
            ])
        
        # Calculate regular vs overtime (assuming 40 hours/week)
        regular_hours = min(total_hours, 40.0)
        overtime_hours = max(0, total_hours - 40.0)
        
        # Add to summary sheet
        summary_ws.append([
            employee.name,
            f"{total_hours:.2f}",
            f"{regular_hours:.2f}",
            f"{overtime_hours:.2f}",
            total_break_minutes,
        ])
    
    # Auto-adjust column widths
    for ws in [summary_ws, detail_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


async def generate_payroll_pdf(
    db: AsyncSession,
    payroll_run,
) -> BytesIO:
    """Generate professional PDF payroll report."""
    from app.models.payroll import PayrollRun, PayrollLineItem
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                           leftMargin=0.5*inch, rightMargin=0.5*inch,
                           topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()
    
    # Professional color scheme
    primary_color = colors.HexColor('#2563eb')
    secondary_color = colors.HexColor('#1e40af')
    header_bg = colors.HexColor('#1e293b')
    light_bg = colors.HexColor('#f8fafc')
    border_color = colors.HexColor('#e2e8f0')
    
    # Simplified header - smaller and cleaner
    company_name = payroll_run.company.name if payroll_run.company else "Company"
    story.append(Spacer(1, 0.1 * inch))
    
    # Company name and title on same line, smaller
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Normal'],
        fontSize=14,
        textColor=header_bg,
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
    )
    
    story.append(Paragraph(f"<b>{company_name}</b> - Payroll Report", header_style))
    
    # Period info only - one line, smaller
    period_text = f"{payroll_run.period_start_date.strftime('%b %d, %Y')} to {payroll_run.period_end_date.strftime('%b %d, %Y')}"
    story.append(Paragraph(period_text, ParagraphStyle(
        'Period',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=15,
        alignment=TA_CENTER,
    )))
    
    # Table data
    data = [["Employee", "Regular Hours", "OT Hours", "Rate", "Regular Pay", "OT Pay", "Total Pay", "Exceptions"]]
    
    for item in payroll_run.line_items:
        regular_hours = item.regular_minutes / 60.0
        ot_hours = item.overtime_minutes / 60.0
        rate_dollars = item.pay_rate_cents / 100.0
        regular_pay_dollars = item.regular_pay_cents / 100.0
        ot_pay_dollars = item.overtime_pay_cents / 100.0
        total_pay_dollars = item.total_pay_cents / 100.0
        
        data.append([
            item.employee.name if item.employee else "Unknown",
            f"{regular_hours:.2f}",
            f"{ot_hours:.2f}",
            f"${rate_dollars:.2f}",
            f"${regular_pay_dollars:.2f}",
            f"${ot_pay_dollars:.2f}",
            f"${total_pay_dollars:.2f}",
            str(item.exceptions_count) if item.exceptions_count > 0 else "-",
        ])
    
    # Totals row - use Paragraph objects with white text for proper bold formatting
    total_regular_hours = float(payroll_run.total_regular_hours)
    total_ot_hours = float(payroll_run.total_overtime_hours)
    total_gross_dollars = payroll_run.total_gross_pay_cents / 100.0
    
    totals_style = ParagraphStyle(
        'Totals',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica-Bold',
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    
    # Convert totals row to use Paragraph for proper rendering with white text
    totals_row = [
        Paragraph("TOTALS", totals_style),
        Paragraph(f"{total_regular_hours:.2f}", totals_style),
        Paragraph(f"{total_ot_hours:.2f}", totals_style),
        "",
        "",
        "",
        Paragraph(f"${total_gross_dollars:,.2f}", totals_style),
        "",
    ]
    
    # Convert all previous rows to Paragraphs for consistency
    table_data = []
    for i, row in enumerate(data):
        if i == 0:  # Header row - white text
            header_style = ParagraphStyle(
                'Header',
                parent=styles['Normal'],
                fontSize=9,
                fontName='Helvetica-Bold',
                textColor=colors.white,
                alignment=TA_CENTER,
            )
            table_data.append([Paragraph(str(cell), header_style) for cell in row])
        else:
            cell_style = ParagraphStyle(
                'Cell',
                parent=styles['Normal'],
                fontSize=8,
                alignment=TA_CENTER,
            )
            table_data.append([Paragraph(str(cell), cell_style) if cell else "" for cell in row])
    
    # Add totals row
    table_data.append(totals_row)
    data = table_data
    
    # Create table with compact styling - smaller fonts and padding
    table = Table(data, colWidths=[2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.1*inch, 1.1*inch, 1.1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        # Header - smaller
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        
        # Data rows - smaller
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -2), colors.HexColor('#1e293b')),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, light_bg]),
        
        # Totals row - smaller
        ('BACKGROUND', (0, -1), (-1, -1), secondary_color),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 9),
        ('TOPPADDING', (0, -1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
        
        # Grid - lighter
        ('GRID', (0, 0), (-1, -1), 0.5, border_color),
        ('LINEBELOW', (0, 0), (-1, 0), 1, header_bg),
        ('LINEABOVE', (0, -1), (-1, -1), 1, secondary_color),
        
        # Padding - reduced
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    
    # No page numbering to keep it simple
    doc.build(story)
    buffer.seek(0)
    return buffer


async def generate_payroll_excel(
    db: AsyncSession,
    payroll_run,
) -> BytesIO:
    """Generate Excel payroll report."""
    from app.models.payroll import PayrollRun, PayrollLineItem
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Summary sheet
    summary_ws = wb.create_sheet("Payroll Summary")
    summary_ws.append(["Payroll Report"])
    summary_ws.append([])
    summary_ws.append(["Company:", payroll_run.company.name if payroll_run.company else "Company"])
    summary_ws.append(["Payroll Type:", payroll_run.payroll_type.value])
    summary_ws.append(["Period Start:", payroll_run.period_start_date.isoformat()])
    summary_ws.append(["Period End:", payroll_run.period_end_date.isoformat()])
    summary_ws.append(["Generated:", payroll_run.generated_at.strftime('%Y-%m-%d %H:%M:%S')])
    if payroll_run.generator:
        summary_ws.append(["Generated By:", payroll_run.generator.name])
    summary_ws.append(["Status:", payroll_run.status.value])
    summary_ws.append([])
    
    # Summary table headers
    summary_ws.append([
        "Employee Name",
        "Regular Hours",
        "Overtime Hours",
        "Pay Rate",
        "Regular Pay",
        "OT Pay",
        "Total Pay",
        "Exceptions",
    ])
    
    # Summary data
    for item in payroll_run.line_items:
        regular_hours = item.regular_minutes / 60.0
        ot_hours = item.overtime_minutes / 60.0
        rate_dollars = item.pay_rate_cents / 100.0
        regular_pay_dollars = item.regular_pay_cents / 100.0
        ot_pay_dollars = item.overtime_pay_cents / 100.0
        total_pay_dollars = item.total_pay_cents / 100.0
        
        summary_ws.append([
            item.employee.name if item.employee else "Unknown",
            regular_hours,
            ot_hours,
            rate_dollars,
            regular_pay_dollars,
            ot_pay_dollars,
            total_pay_dollars,
            item.exceptions_count,
        ])
    
    # Totals row
    total_regular_hours = float(payroll_run.total_regular_hours)
    total_ot_hours = float(payroll_run.total_overtime_hours)
    total_gross_dollars = payroll_run.total_gross_pay_cents / 100.0
    
    summary_ws.append([
        "TOTALS",
        total_regular_hours,
        total_ot_hours,
        "",
        "",
        "",
        total_gross_dollars,
        "",
    ])
    
    # Details sheet
    detail_ws = wb.create_sheet("Payroll Details")
    detail_ws.append(["Payroll Details"])
    detail_ws.append([])
    detail_ws.append([
        "Employee",
        "Date",
        "Minutes",
        "Regular Minutes",
        "OT Minutes",
        "Notes/Exceptions",
    ])
    
    for item in payroll_run.line_items:
        employee_name = item.employee.name if item.employee else "Unknown"
        details = item.details_json or {}
        days = details.get("days", {})
        
        for date_str, minutes in days.items():
            # Determine if minutes are regular or OT (simplified - would need week breakdown)
            # For now, just show total minutes
            detail_ws.append([
                employee_name,
                date_str,
                minutes,
                "",  # Would need to calculate from week_blocks
                "",  # Would need to calculate from week_blocks
                f"Exceptions: {item.exceptions_count}" if item.exceptions_count > 0 else "",
            ])
    
    # Style headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for ws in [summary_ws, detail_ws]:
        for row in ws.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Auto-adjust column widths
    for ws in [summary_ws, detail_ws]:
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
